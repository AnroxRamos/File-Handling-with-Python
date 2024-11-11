from openpyxl import Workbook, load_workbook

# Writing to an Excel file
wb = Workbook()
ws = wb.active
ws['A1'] = "ID"
ws['B1'] = "Name"
ws.append([1, "John"])
wb.save("data.xlsx")

# Reading an Excel file
wb = load_workbook("data.xlsx")
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
